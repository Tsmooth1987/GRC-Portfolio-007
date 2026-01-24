#!/usr/bin/env python3
"""
Create a Shareable AWS Audit Summary

This script creates a simplified, clean Excel report from the detailed audit.
"""
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
from datetime import datetime

def create_shareable_report(audit_file, output_file):
    """Create a simplified, shareable Excel report."""
    # Read the audit data
    xls = pd.ExcelFile(audit_file)
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Summary"
    
    # Styles
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    highlight_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"AWS Security Audit Summary - {datetime.now().strftime('%B %d, %Y')}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    
    # Summary Section
    ws['A3'] = "Account ID:"
    ws['B3'] = "291846646255"
    ws['A4'] = "Report Date:"
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Key Metrics
    metrics = [
        ["IAM Users", len(pd.read_excel(xls, 'IAM Users')) if 'IAM Users' in xls.sheet_names else 0],
        ["S3 Buckets", len(pd.read_excel(xls, 'S3 Buckets')) if 'S3 Buckets' in xls.sheet_names else 0],
        ["Security Groups", len(pd.read_excel(xls, 'Security Groups')) if 'Security Groups' in xls.sheet_names else 0],
        ["EC2 Instances", len(pd.read_excel(xls, 'EC2 Instances')) if 'EC2 Instances' in xls.sheet_names else 0]
    ]
    
    ws['A6'] = "Key Metrics"
    ws['A6'].font = Font(bold=True)
    
    for i, (metric, value) in enumerate(metrics, start=7):
        ws[f'A{i}'] = metric
        ws[f'B{i}'] = value
    
    # Critical Findings
    if 'Audit Findings' in xls.sheet_names:
        findings = pd.read_excel(xls, 'Audit Findings')
        
        # Group by severity
        severity_counts = findings['Severity'].value_counts().to_dict()
        
        ws['A12'] = "Critical Findings by Severity"
        ws['A12'].font = Font(bold=True)
        
        for i, (severity, count) in enumerate(severity_counts.items(), start=13):
            ws[f'A{i}'] = f"{severity} Severity"
            ws[f'B{i}'] = count
    
    # Action Items
    ws['A18'] = "Recommended Actions"
    ws['A18'].font = Font(bold=True)
    
    actions = [
        "Review and enable MFA for all IAM users",
        "Check public S3 buckets and restrict access if needed",
        "Review and harden security group rules",
        "Review IAM policies for least privilege access",
        "Schedule monthly security audits"
    ]
    
    for i, action in enumerate(actions, start=19):
        ws[f'A{i}'] = f"{i-18}."
        ws[f'B{i}'] = action
    
    # Security Score
    ws['A25'] = "Security Score"
    ws['A25'].font = Font(size=12, bold=True)
    
    # Calculate a simple security score (0-100)
    total_checks = 0
    passed_checks = 0
    
    # Check 1: MFA enabled for all users
    if 'IAM Users' in xls.sheet_names:
        iam_users = pd.read_excel(xls, 'IAM Users')
        total_users = len(iam_users)
        if total_users > 0:
            total_checks += 1
            mfa_enabled = sum(1 for _, user in iam_users.iterrows() if user.get('MFAEnabled') == 'Yes')
            if mfa_enabled == total_users:
                passed_checks += 1
    
    # Check 2: No public S3 buckets
    if 'S3 Buckets' in xls.sheet_names:
        s3_buckets = pd.read_excel(xls, 'S3 Buckets')
        total_buckets = len(s3_buckets)
        if total_buckets > 0:
            total_checks += 1
            public_buckets = sum(1 for _, bucket in s3_buckets.iterrows() if bucket.get('IsPublic') == 'Yes')
            if public_buckets == 0:
                passed_checks += 1
    
    # Calculate score (avoid division by zero)
    security_score = (passed_checks / total_checks * 100) if total_checks > 0 else 0
    
    # Add score to worksheet
    ws['A26'] = "Current Security Score:"
    ws['B26'] = f"{security_score:.1f}%"
    ws['B26'].font = Font(size=14, bold=True, color='4F81BD')
    
    # Add a simple gauge visualization
    ws['A28'] = "Security Status:"
    if security_score >= 80:
        ws['B28'] = "🟢 Secure"
    elif security_score >= 50:
        ws['B28'] = "🟡 Needs Improvement"
    else:
        ws['B28'] = "🔴 Requires Immediate Attention"
    ws['B28'].font = Font(bold=True)
    
    # Formatting
    for row in ws.iter_rows():
        for cell in row:
            # Skip merged cells
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
                
            cell.border = thin_border
            if cell.row == 1:
                cell.font = Font(size=14, bold=True)
            elif cell.row in [6, 12, 18]:
                cell.font = Font(bold=True)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = None
        
        # Get the column letter from the first non-merged cell
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                column_letter = cell.column_letter
                break
                
        if column_letter is None:
            continue
            
        # Find the maximum length of content in the column
        for cell in col:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
                
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
                
        # Set column width (with some padding)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 50)
    
    # Save the workbook
    wb.save(output_file)
    print(f"Shareable report created: {output_file}")

if __name__ == "__main__":
    import sys
    
    # Find the latest audit file
    import glob
    audit_files = sorted(glob.glob('aws_audit_report_*.xlsx'), reverse=True)
    
    if not audit_files:
        print("No audit files found. Please run the audit first.")
        sys.exit(1)
    
    latest_audit = audit_files[0]
    output_file = "AWS_Security_Summary.xlsx"
    
    create_shareable_report(latest_audit, output_file)
