#!/usr/bin/env python3
"""
AWS Security Audit Report Generator

This script connects to AWS, collects security-related information,
and generates an Excel report with audit-ready formatting.
"""
import boto3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from botocore.exceptions import ClientError
import time
from typing import Dict, List, Any, Optional

def get_aws_session():
    """Initialize and return AWS session."""
    try:
        # Using the 'tj' profile from AWS credentials
        session = boto3.Session(profile_name='tj')
        
        # Verify the session works by getting the caller identity
        sts = session.client('sts')
        identity = sts.get_caller_identity()
        print(f"Successfully authenticated to AWS Account: {identity['Account']} as {identity['Arn']}")
        
        return session
    except Exception as e:
        print(f"Error creating AWS session: {e}")
        print("\nTroubleshooting steps:")
        print("1. Ensure your AWS credentials in ~/.aws/credentials are correct")
        print("2. Verify the 'tj' profile has the necessary IAM permissions")
        print("3. Check your internet connection and AWS service health")
        return None

def get_iam_users(session) -> List[Dict[str, Any]]:
    """Retrieve detailed IAM user information."""
    iam = session.client('iam')
    users = []
    try:
        paginator = iam.get_paginator('list_users')
        for page in paginator.paginate():
            for user in page['Users']:
                # Get user details
                user_details = {
                    'UserName': user['UserName'],
                    'UserId': user['UserId'],
                    'CreateDate': user['CreateDate'].strftime('%Y-%m-%d %H:%M:%S'),
                    'PasswordLastUsed': user.get('PasswordLastUsed', 'Never').strftime('%Y-%m-%d %H:%M:%S') 
                                    if 'PasswordLastUsed' in user else 'Never',
                    'MFAEnabled': 'No'
                }
                
                # Check for MFA devices
                try:
                    mfa_devices = iam.list_mfa_devices(UserName=user['UserName'])
                    user_details['MFAEnabled'] = 'Yes' if mfa_devices['MFADevices'] else 'No'
                except ClientError as e:
                    user_details['MFAEnabled'] = 'Error checking'
                
                users.append(user_details)
    except ClientError as e:
        print(f"Error fetching IAM users: {e}")
    return users

def get_ec2_instances(session) -> List[Dict[str, Any]]:
    """Retrieve detailed EC2 instance information."""
    ec2 = session.client('ec2')
    instances = []
    
    try:
        # Get all regions
        regions = [region['RegionName'] for region in ec2.describe_regions()['Regions']]
        
        for region in regions:
            try:
                ec2_region = session.client('ec2', region_name=region)
                response = ec2_region.describe_instances()
                
                for reservation in response['Reservations']:
                    for instance in reservation['Instances']:
                        # Get instance name from tags
                        name = next((tag['Value'] for tag in instance.get('Tags', [])
                                  if tag['Key'] == 'Name'), 'N/A')
                        
                        # Get security groups
                        security_groups = ", ".join([sg['GroupName'] for sg in instance.get('SecurityGroups', [])])
                        
                        # Get public/private IPs
                        public_ip = instance.get('PublicIpAddress', 'N/A')
                        private_ip = instance.get('PrivateIpAddress', 'N/A')
                        
                        instances.append({
                            'InstanceId': instance['InstanceId'],
                            'Name': name,
                            'Type': instance['InstanceType'],
                            'State': instance['State']['Name'],
                            'LaunchTime': instance['LaunchTime'].strftime('%Y-%m-%d %H:%M:%S'),
                            'Region': region,
                            'PublicIP': public_ip,
                            'PrivateIP': private_ip,
                            'SecurityGroups': security_groups,
                            'VPC': instance.get('VpcId', 'N/A'),
                            'Subnet': instance.get('SubnetId', 'N/A')
                        })
            except ClientError as e:
                print(f"Error in region {region}: {e}")
                continue
                
    except ClientError as e:
        print(f"Error fetching EC2 instances: {e}")
    
    return instances

def get_s3_buckets(session) -> List[Dict[str, Any]]:
    """Retrieve S3 bucket information."""
    s3 = session.client('s3')
    s3_resource = session.resource('s3')
    buckets = []
    
    try:
        response = s3.list_buckets()
        for bucket in response['Buckets']:
            try:
                # Get bucket location (region)
                location = s3.get_bucket_location(Bucket=bucket['Name'])['LocationConstraint']
                # us-east-1 returns None, so we set it explicitly
                location = location if location else 'us-east-1'
                
                # Get bucket size and object count
                total_size = 0
                total_objects = 0
                
                try:
                    for obj in s3_resource.Bucket(bucket['Name']).objects.all():
                        total_size += obj.size
                        total_objects += 1
                except Exception as e:
                    print(f"Could not get size for bucket {bucket['Name']}: {e}")
                
                # Get bucket policy status (public/private)
                is_public = False
                try:
                    acl = s3.get_bucket_acl(Bucket=bucket['Name'])
                    for grant in acl.get('Grants', []):
                        if 'URI' in grant.get('Grantee', {}) and 'AllUsers' in grant['Grantee']['URI']:
                            is_public = True
                            break
                except Exception as e:
                    print(f"Could not check bucket policy for {bucket['Name']}: {e}")
                
                buckets.append({
                    'Name': bucket['Name'],
                    'CreationDate': bucket['CreationDate'].strftime('%Y-%m-%d %H:%M:%S'),
                    'Region': location,
                    'Size (MB)': round(total_size / (1024 * 1024), 2),
                    'ObjectCount': total_objects,
                    'IsPublic': 'Yes' if is_public else 'No'
                })
                
            except ClientError as e:
                print(f"Error processing bucket {bucket.get('Name', 'unknown')}: {e}")
                continue
                
    except ClientError as e:
        print(f"Error listing S3 buckets: {e}")
    
    return buckets

def get_security_groups(session) -> List[Dict[str, Any]]:
    """Retrieve security group information."""
    ec2 = session.client('ec2')
    security_groups = []
    
    try:
        # Get all regions
        regions = [region['RegionName'] for region in ec2.describe_regions()['Regions']]
        
        for region in regions:
            try:
                ec2_region = session.client('ec2', region_name=region)
                response = ec2_region.describe_security_groups()
                
                for sg in response['SecurityGroups']:
                    # Count inbound and outbound rules
                    inbound_rules = len(sg.get('IpPermissions', []))
                    outbound_rules = len(sg.get('IpPermissionsEgress', []))
                    
                    # Check for overly permissive rules
                    is_permissive = False
                    for perm in sg.get('IpPermissions', []):
                        for ip_range in perm.get('IpRanges', []):
                            if ip_range.get('CidrIp') == '0.0.0.0/0':
                                is_permissive = True
                                break
                    
                    security_groups.append({
                        'GroupId': sg['GroupId'],
                        'GroupName': sg['GroupName'],
                        'Description': sg.get('Description', ''),
                        'VPC': sg.get('VpcId', 'N/A'),
                        'Region': region,
                        'InboundRules': inbound_rules,
                        'OutboundRules': outbound_rules,
                        'IsPermissive': 'Yes' if is_permissive else 'No',
                        'AssociatedInstances': len(sg.get('IpPermissionsEgress', []))  # Simplified count
                    })
                    
            except ClientError as e:
                print(f"Error in region {region}: {e}")
                continue
                
    except ClientError as e:
        print(f"Error fetching security groups: {e}")
    
    return security_groups

def format_excel_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    """Apply formatting to Excel sheet."""
    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Define styles
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styles
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        
        # Auto-adjust column width
        column_letter = get_column_letter(col_num)
        max_length = max(df[column_title].astype(str).apply(len).max(), len(column_title)) + 2
        worksheet.column_dimensions[column_letter].width = min(max_length, 30)
    
    # Apply cell styles
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                 min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'
    
    # Add filter to header
    worksheet.auto_filter.ref = worksheet.dimensions

def generate_excel_report(users: List[Dict[str, Any]], 
                        instances: List[Dict[str, Any]],
                        buckets: List[Dict[str, Any]],
                        security_groups: List[Dict[str, Any]],
                        output_file: str = 'aws_audit_report.xlsx') -> bool:
    """Generate Excel report from collected data with enhanced formatting."""
    try:
        # Create a timestamp for the report
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # 1. Summary Sheet
            summary_data = {
                'Metric': [
                    'Report Generated',
                    'Total IAM Users',
                    'Total EC2 Instances',
                    'Total S3 Buckets',
                    'Total Security Groups',
                    'Public S3 Buckets',
                    'EC2 Instances with Public IPs'
                ],
                'Value': [
                    timestamp,
                    len(users),
                    len(instances),
                    len(buckets),
                    len(security_groups),
                    len([b for b in buckets if b.get('IsPublic') == 'Yes']),
                    len([i for i in instances if i.get('PublicIP') != 'N/A'])
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # 2. IAM Users Sheet
            if users:
                df_users = pd.DataFrame(users)
                df_users.to_excel(writer, sheet_name='IAM Users', index=False)
                format_excel_sheet(writer, 'IAM Users', df_users)
            
            # 3. EC2 Instances Sheet
            if instances:
                df_instances = pd.DataFrame(instances)
                df_instances.to_excel(writer, sheet_name='EC2 Instances', index=False)
                format_excel_sheet(writer, 'EC2 Instances', df_instances)
            
            # 4. S3 Buckets Sheet
            if buckets:
                df_buckets = pd.DataFrame(buckets)
                df_buckets.to_excel(writer, sheet_name='S3 Buckets', index=False)
                format_excel_sheet(writer, 'S3 Buckets', df_buckets)
            
            # 5. Security Groups Sheet
            if security_groups:
                df_sg = pd.DataFrame(security_groups)
                df_sg.to_excel(writer, sheet_name='Security Groups', index=False)
                format_excel_sheet(writer, 'Security Groups', df_sg)
            
            # 6. Audit Findings Sheet
            findings = []
            
            # IAM findings
            for user in users:
                if user.get('MFAEnabled') != 'Yes':
                    findings.append({
                        'Severity': 'High',
                        'Service': 'IAM',
                        'Resource': user['UserName'],
                        'Finding': 'MFA not enabled for IAM user',
                        'Recommendation': 'Enable MFA for all IAM users'
                    })
                
                if user.get('PasswordLastUsed') == 'Never':
                    findings.append({
                        'Severity': 'Medium',
                        'Service': 'IAM',
                        'Resource': user['UserName'],
                        'Finding': 'IAM user has never used password',
                        'Recommendation': 'Remove or investigate unused IAM users'
                    })
            
            # EC2 findings
            for instance in instances:
                if instance.get('PublicIP') != 'N/A' and not instance.get('SecurityGroups'):
                    findings.append({
                        'Severity': 'High',
                        'Service': 'EC2',
                        'Resource': instance['InstanceId'],
                        'Finding': 'EC2 instance has a public IP but no security groups',
                        'Recommendation': 'Apply appropriate security groups or move to private subnet'
                    })
            
            # S3 findings
            for bucket in buckets:
                if bucket.get('IsPublic') == 'Yes':
                    findings.append({
                        'Severity': 'High',
                        'Service': 'S3',
                        'Resource': bucket['Name'],
                        'Finding': 'S3 bucket is publicly accessible',
                        'Recommendation': 'Review and update bucket policy to restrict public access'
                    })
            
            # Security Group findings
            for sg in security_groups:
                if sg.get('IsPermissive') == 'Yes':
                    findings.append({
                        'Severity': 'High',
                        'Service': 'EC2',
                        'Resource': sg['GroupId'],
                        'Finding': 'Security group has overly permissive rules',
                        'Recommendation': 'Review and restrict security group rules'
                    })
            
            if findings:
                df_findings = pd.DataFrame(findings)
                df_findings.to_excel(writer, sheet_name='Audit Findings', index=False)
                format_excel_sheet(writer, 'Audit Findings', df_findings)
            
            # Format the Summary sheet last (after it's been created)
            format_excel_sheet(writer, 'Summary', df_summary)
            
        print(f"\n✅ Report generated successfully: {output_file}")
        print(f"📊 Sheets included: Summary, IAM Users, EC2 Instances, S3 Buckets, Security Groups, Audit Findings")
        return True
        
    except Exception as e:
        print(f"\n❌ Error generating Excel report: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    print("🚀 Starting AWS Security Audit Report Generation...\n")
    
    # Initialize AWS session
    print("🔑 Initializing AWS session...")
    session = get_aws_session()
    if not session:
        print("❌ Failed to initialize AWS session. Please check your AWS credentials.")
        return
    
    print("✅ AWS session established successfully\n")
    
    # Collect data with progress indicators
    print("📋 Collecting IAM users...")
    users = get_iam_users(session)
    print(f"   → Found {len(users)} IAM users")
    
    print("\n🖥️  Collecting EC2 instances across all regions...")
    instances = get_ec2_instances(session)
    print(f"   → Found {len(instances)} EC2 instances")
    
    print("\n📦 Collecting S3 bucket information...")
    buckets = get_s3_buckets(session)
    print(f"   → Found {len(buckets)} S3 buckets")
    
    print("\n🛡️  Collecting security group information...")
    security_groups = get_security_groups(session)
    print(f"   → Found {len(security_groups)} security groups")
    
    # Generate report
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'aws_audit_report_{timestamp}.xlsx'
    
    print("\n📊 Generating Excel report...")
    if generate_excel_report(users, instances, buckets, security_groups, output_file):
        print("\n✨ Audit report generation complete!")
        print("\nNext steps:")
        print(f"1. 📂 Review the generated Excel file: {output_file}")
        print("2. 🔍 Check the 'Audit Findings' sheet for potential security issues")
        print("3. 📧 Share with relevant stakeholders")
        print("4. 🔄 Schedule regular audits (recommended: monthly)")
    else:
        print("\n❌ Failed to generate audit report. Please check the error messages above.")

if __name__ == "__main__":
    main()
