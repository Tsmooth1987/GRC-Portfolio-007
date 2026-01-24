"""
AWS Security Hub to Excel Report Generator
Lambda function that extracts Security Hub findings and generates professional Excel reports

Author: Terence Webster
Project: GRC Portfolio - Project 8
"""

import json
import boto3
import logging
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from botocore.exceptions import ClientError

# Configure logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# AWS clients
s3_client = boto3.client('s3')
securityhub_client = boto3.client('securityhub')


def lambda_handler(event, context):
    """Main Lambda handler function"""
    try:
        logger.info("Starting Security Hub to Excel report generation")
        
        import os
        bucket_name = os.environ.get('S3_BUCKET_NAME')
        if not bucket_name:
            raise ValueError("S3_BUCKET_NAME environment variable not set")
        
        findings = fetch_security_hub_findings()
        logger.info(f"Retrieved {len(findings)} Security Hub findings")
        
        if not findings:
            logger.warning("No findings found in Security Hub")
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'message': 'No Security Hub findings to report',
                    'findings_count': 0
                })
            }
        
        excel_buffer = generate_excel_report(findings)
        
        report_key = f"reports/security_hub_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        s3_client.put_object(
            Bucket=bucket_name,
            Key=report_key,
            Body=excel_buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        logger.info(f"Report uploaded successfully to s3://{bucket_name}/{report_key}")
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Report generated successfully',
                'findings_count': len(findings),
                's3_bucket': bucket_name,
                's3_key': report_key,
                's3_url': f"s3://{bucket_name}/{report_key}"
            })
        }
        
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}", exc_info=True)
        return {
            'statusCode': 500,
            'body': json.dumps({
                'error': str(e),
                'message': 'Failed to generate Security Hub report'
            })
        }


def fetch_security_hub_findings():
    """Fetch findings from AWS Security Hub"""
    findings = []
    
    try:
        paginator = securityhub_client.get_paginator('get_findings')
        
        filters = {
            'WorkflowStatus': [{'Value': 'NEW', 'Comparison': 'EQUALS'}],
            'RecordState': [{'Value': 'ACTIVE', 'Comparison': 'EQUALS'}]
        }
        
        for page in paginator.paginate(Filters=filters, MaxResults=100):
            findings.extend(page['Findings'])
            
        logger.info(f"Successfully fetched {len(findings)} findings from Security Hub")
        
    except ClientError as e:
        logger.error(f"Error fetching Security Hub findings: {e}")
        raise
    
    return findings


def generate_excel_report(findings):
    """Generate a professional Excel report from Security Hub findings"""
    wb = Workbook()
    
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_details = wb.create_sheet("Detailed Findings")
    ws_critical = wb.create_sheet("Critical & High")
    
    create_summary_sheet(ws_summary, findings)
    create_details_sheet(ws_details, findings)
    create_critical_sheet(ws_critical, findings)
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def create_summary_sheet(ws, findings):
    """Create executive summary worksheet with key metrics"""
    ws['A1'] = 'AWS Security Hub - Executive Summary'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    ws['A3'] = 'Report Generated:'
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')
    ws['A4'] = 'Total Findings:'
    ws['B4'] = len(findings)
    
    severity_counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFORMATIONAL': 0}
    
    for finding in findings:
        severity = finding.get('Severity', {}).get('Label', 'INFORMATIONAL')
        if severity in severity_counts:
            severity_counts[severity] += 1
    
    ws['A6'] = 'Severity Breakdown'
    ws['A6'].font = Font(bold=True, size=12)
    
    row = 7
    headers = ['Severity', 'Count', 'Percentage']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    severity_colors = {
        'CRITICAL': 'C00000', 'HIGH': 'FF6B6B', 'MEDIUM': 'FFA500',
        'LOW': 'FFD966', 'INFORMATIONAL': '92D050'
    }
    
    for severity, count in severity_counts.items():
        percentage = (count / len(findings) * 100) if findings else 0
        ws.cell(row=row, column=1, value=severity)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
        
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=severity_colors[severity], 
                                   end_color=severity_colors[severity], fill_type='solid')
            if severity in ['CRITICAL', 'HIGH']:
                cell.font = Font(color='FFFFFF', bold=True)
        row += 1
    
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_details_sheet(ws, findings):
    """Create detailed findings worksheet"""
    headers = ['Finding ID', 'Title', 'Severity', 'Compliance Status', 'Resource Type',
               'Resource ID', 'AWS Account', 'Region', 'First Observed', 'Last Observed',
               'Description', 'Remediation']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row, finding in enumerate(findings, start=2):
        resources = finding.get('Resources', [])
        resource_type = resources[0].get('Type', 'N/A') if resources else 'N/A'
        resource_id = resources[0].get('Id', 'N/A') if resources else 'N/A'
        remediation = finding.get('Remediation', {}).get('Recommendation', {}).get('Text', 'N/A')
        
        data = [
            finding.get('Id', 'N/A'),
            finding.get('Title', 'N/A'),
            finding.get('Severity', {}).get('Label', 'N/A'),
            finding.get('Compliance', {}).get('Status', 'N/A'),
            resource_type, resource_id,
            finding.get('AwsAccountId', 'N/A'),
            finding.get('Region', 'N/A'),
            finding.get('FirstObservedAt', 'N/A'),
            finding.get('LastObservedAt', 'N/A'),
            finding.get('Description', 'N/A'),
            remediation
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=str(value))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            if col == 3:
                severity = value
                if severity == 'CRITICAL':
                    cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
                elif severity == 'HIGH':
                    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                    cell.font = Font(bold=True)
    
    for col in range(1, len(headers) + 1):
        if col in [11, 12]:
            ws.column_dimensions[get_column_letter(col)].width = 50
        else:
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    ws.freeze_panes = 'A2'


def create_critical_sheet(ws, findings):
    """Create worksheet with only CRITICAL and HIGH severity findings"""
    critical_findings = [f for f in findings 
                        if f.get('Severity', {}).get('Label', '') in ['CRITICAL', 'HIGH']]
    
    if not critical_findings:
        ws['A1'] = 'No Critical or High Severity Findings'
        ws['A1'].font = Font(size=14, bold=True)
        return
    
    headers = ['Severity', 'Title', 'Resource Type', 'Resource ID',
               'Compliance Status', 'Description', 'Remediation Steps']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row, finding in enumerate(critical_findings, start=2):
        resources = finding.get('Resources', [])
        resource_type = resources[0].get('Type', 'N/A') if resources else 'N/A'
        resource_id = resources[0].get('Id', 'N/A') if resources else 'N/A'
        remediation = finding.get('Remediation', {}).get('Recommendation', {}).get('Text', 'N/A')
        
        data = [
            finding.get('Severity', {}).get('Label', 'N/A'),
            finding.get('Title', 'N/A'),
            resource_type, resource_id,
            finding.get('Compliance', {}).get('Status', 'N/A'),
            finding.get('Description', 'N/A'),
            remediation
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=str(value))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            severity = finding.get('Severity', {}).get('Label', '')
            if severity == 'CRITICAL':
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            elif severity == 'HIGH':
                cell.fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
    
    for col in range(1, len(headers) + 1):
        if col in [6, 7]:
            ws.column_dimensions[get_column_letter(col)].width = 50
        else:
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    ws.freeze_panes = 'A2'
